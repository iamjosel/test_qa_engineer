import pandas as pd
from openpyxl import Workbook

# Ruta para el reporte Excel
report_path = "O:\\test_qa_engineer\\test_report.xlsx"

df = "O:\\test_qa_engineer\\ventas1.xlsx"
data = pd.read_excel(df)

# Preparación del reporte
wb = Workbook()
ws = wb.active
ws.title = "Resultados de casos de prueba"
# Información del tester y fecha de ejecución
ws['A1'] = "Nombre del tester: José Luis García Quinayás"
ws['A2'] = "Fecha de ejecución de pruebas: 27/Oct/2024"
ws['A3'] = "Versión 1.0"
ws['A4'] = ""
ws.append(["ID Caso", "Descripción Caso de Prueba", "Resultado", "Descripción del Resultado", "Número de Incidentes", "Observaciones"])

# Cargar dataset
def load_data():
    try:
        data = pd.read_excel("O:\\jose-test\\ventas1.xlsx")
        # Convertir columnas a numérico según sea necesario
        data['precio'] = pd.to_numeric(data['precio'], errors='coerce')
        data['cantidad_vendida'] = pd.to_numeric(data['cantidad_vendida'], errors='coerce')
        data['total_venta'] = pd.to_numeric(data['total_venta'], errors='coerce')
        return data
    except Exception as e:
        print(f"Error loading data: {e}")

data = load_data()

# Función para registrar resultados en Excel
def record_result(test_id, description, result, result_description, incidents, comments):
    ws.append([test_id, description, result, result_description, incidents, comments])

### Funciones de Validación

# 1. Validar formato de fecha (YYYY-MM-DD)
def validate_date_format(data):
    test_id = "Caso 01"
    description = "Validar formato de fecha del campo'fecha_venta' (AAAA-MM-DD)"
    try:
        incorrect_dates = data[~data['fecha_venta'].astype(str).str.match(r'^\d{4}-\d{2}-\d{2}$')]
        result = "Aprobado" if incorrect_dates.empty else "Fallido"
        record_result(test_id, description, result, f"Fechas no validas: {len(incorrect_dates)}", len(incorrect_dates), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Error de validación del formato de fecha")

# 2. Validar valores numéricos positivos
def validate_positive_numbers(data):
    test_id = "Caso 02"
    description = "Validar números positivos para los campos 'id_producto', 'precio', 'cantidad_vendida', 'total_venta'"
    try:
        incorrect_values = data[(data[['precio', 'cantidad_vendida', 'total_venta']] <= 0).any(axis=1)]
        result = "Aprobado" if incorrect_values.empty else "Fallido"
        record_result(test_id, description, result, f"Valores negativos o cero: {len(incorrect_values)}", len(incorrect_values), "")
    except TypeError:
        record_result(test_id, description, "Fallido", "Error encontrado; asegurar valores numéricos", 1, "Conversion requerida a tipo número")

# 3. Verificar campos no numéricos en ciertos campos
def validate_no_numbers_in_strings(data):
    test_id = "Caso 03"
    description = "Verificar que no hay números en 'nombre_cliente', 'nombre_producto', 'categoria', 'region', 'metodo_pago'"
    try:
        invalid_strings = data[(data[['nombre_cliente', 'nombre_producto', 'categoria', 'region', 'metodo_pago']]
                               .apply(lambda x: x.str.contains(r'\d', na=False))).any(axis=1)]
        result = "Aprobado" if invalid_strings.empty else "Fallido"
        record_result(test_id, description, result, f"Campos que contienen números: {len(invalid_strings)}", len(invalid_strings), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Error al validar campos de caracteres")

# 4. Verificar que no haya datos vacíos o nulos
def validate_no_nulls(data):
    test_id = "Caso 04"
    description = "Validar de que no haya valores nulos o vacíos en todos los campos."
    try:
        empty_fields = data[data[['fecha_venta', 'nombre_producto', 'categoria', 'precio', 'cantidad_vendida', 'total_venta', 
                                  'nombre_cliente', 'region', 'metodo_pago']].isnull().any(axis=1)]
        result = "Aprobado" if empty_fields.empty else "Fallido"
        record_result(test_id, description, result, f"Campos nulos/vacíos: {len(empty_fields)}", len(empty_fields), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo for nulls")

# 5. Verificar caracteres especiales en campos string
def validate_no_special_chars(data):
    test_id = "Caso 05"
    description = "Validar que no haya caracteres especiales en los campos de cadena"
    try:
        special_chars = data[(data[['nombre_cliente', 'nombre_producto', 'categoria', 'region', 'metodo_pago']]
                              .apply(lambda x: x.str.contains(r'[^a-zA-Z\s]', na=False))).any(axis=1)]
        result = "Aprobado" if special_chars.empty else "Fallido"
        record_result(test_id, description, result, f"Caracteres especiales encontrados: {len(special_chars)}", len(special_chars), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Error al validar caracteres especiales")

# 6. Validar regiones válidas
def validate_region(data):
    test_id = "Caso 06"
    description = "Validar que la región sólo contiene opciones válidas (Norte, Sur, Este, Oeste, Centro)"
    try:
        invalid_regions = data[~data['region'].isin(['Norte', 'Sur', 'Este', 'Oeste', 'Centro'])]
        result = "Aprobado" if invalid_regions.empty else "Fallido"
        record_result(test_id, description, result, f"Regiones no válidas: {len(invalid_regions)}", len(invalid_regions), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo region")

# 7. Validar opciones de método de pago válidas
def validate_payment_method(data):
    test_id = "Caso 07"
    description = "Valirdar que el campo 'metodo_pago' contiene sólo 'Efectivo' o 'Transferencia Bancaria'"
    try:
        invalid_methods = data[~data['metodo_pago'].isin(['Efectivo', 'Transferencia Bancaria'])]
        result = "Aprobado" if invalid_methods.empty else "Fallido"
        record_result(test_id, description, result, f"Métodos de pago no válidos: {len(invalid_methods)}", len(invalid_methods), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Error de validación del método de pago")
# 8. Validar opciones acentuadas en metodo_pago
def validate_accented_payment_method(data):
    test_id = "Caso 08"
    description = "Validar que el campo 'metodo_pago' no contiene palabras acentuadas"
    try:
        invalid_entries = data[data['metodo_pago'].str.contains(r'(?<!Transferencia) Bancaria', regex=True, na=False)]
        result = "Aprobado" if invalid_entries.empty else "Fallido"
        record_result(test_id, description, result, f"Problemas de acentos: {len(invalid_entries)}", len(invalid_entries), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo payment method")

# 10. Validar que 'nombre_cliente' no contenga números o caracteres especiales
def validate_cliente_no_numbers_specials(data):
    test_id = "Caso 09"
    description = "Validar que 'nombre_cliente' no contiene números ni caracteres especiales"
    try:
        invalid_entries = data[data['nombre_cliente'].str.contains(r'[0-9]|[^\w\s]', regex=True, na=False)]
        result = "Aprobado" if invalid_entries.empty else "Fallido"
        record_result(test_id, description, result, f"Los 'nombre_cliente' no válidos: {len(invalid_entries)}", len(invalid_entries), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo 'nombre_cliente'")

# 11. Validar que 'nombre_producto' no contenga números o caracteres especiales
def validate_producto_no_numbers_specials(data):
    test_id = "Caso 10"
    description = "Validar que 'nombre_producto' no contiene números ni caracteres especiales"
    try:
        invalid_entries = data[data['nombre_producto'].str.contains(r'[0-9]|[^\w\s]', regex=True, na=False)]
        result = "Aprobado" if invalid_entries.empty else "Fallido"
        record_result(test_id, description, result, f"El 'nombre_producto' no válidos: {len(invalid_entries)}", len(invalid_entries), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo 'nombre_producto'")

# 12. Validar que 'categoria' no contenga números o caracteres especiales
def validate_categoria_no_numbers_specials(data):
    test_id = "Caso 11"
    description = "Validar que 'categoria' no contiene números ni caracteres especiales"
    try:
        invalid_entries = data[data['categoria'].str.contains(r'[0-9]|[^\w\s]', regex=True, na=False)]
        result = "Aprobado" if invalid_entries.empty else "Fallido"
        record_result(test_id, description, result, f"Las 'categoria' no válidas: {len(invalid_entries)}", len(invalid_entries), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo 'categoria'")

# 13. Validar que 'metodo_pago' no contenga números o caracteres especiales
def validate_payment_no_numbers_specials(data):
    test_id = "Caso 12"
    description = "Validar que 'metodo_pago' no contiene números ni caracteres especiales"
    try:
        invalid_entries = data[data['metodo_pago'].str.contains(r'[0-9]|[^\w\s]', regex=True, na=False)]
        result = "Aprobado" if invalid_entries.empty else "Fallido"
        record_result(test_id, description, result, f"Los 'metodo_pago' inválidos: {len(invalid_entries)}", len(invalid_entries), "")
    except Exception as e:
        record_result(test_id, description, "Fallido", str(e), 1, "Comprobación de errores del campo 'metodo_pago'")

### Ejecutar las pruebas y generar el reporte
def run_tests():
    validate_date_format(data)
    validate_positive_numbers(data)
    validate_no_numbers_in_strings(data)
    validate_no_nulls(data)
    validate_no_special_chars(data)
    validate_region(data)
    validate_payment_method(data)
    validate_accented_payment_method(data)
    validate_cliente_no_numbers_specials(data)
    validate_producto_no_numbers_specials(data)
    validate_categoria_no_numbers_specials(data)
    validate_payment_no_numbers_specials(data)

    wb.save(report_path)

run_tests()
print(f"Reporte guardado en {report_path}")
